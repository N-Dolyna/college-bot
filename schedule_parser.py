from openpyxl import load_workbook
from typing import List, Dict, Optional, Union
import json

class Subject:
    def __init__(self, subject: str, teacher: str):
        self.subject = subject
        self.teacher = teacher
    
    def __str__(self):
        return f"{self.subject} ({self.teacher})"
    
    def to_dict(self):
        return {
            "subject": self.subject,
            "teacher": self.teacher
        }

class Pair:
    def __init__(self, day: str, pair_num: int, begin: str, end: str, subjects=None):
        self.day = day
        self.pair_num = pair_num
        self.begin = begin
        self.end = end
        self.subjects = subjects
    
    def __str__(self):
        base = f"{self.day}, пара {self.pair_num} ({self.begin}-{self.end})"
        if self.subjects is None:
            return f"{base}: - вільно"
        if isinstance(self.subjects, Subject):
            return f"{base}: {self.subjects}"
        if isinstance(self.subjects, list):
            texts = [str(s) for s in self.subjects]
            return f"{base}: {' / '.join(texts)}"
        return f"{base}: {self.subjects}"
    
    def to_dict(self):
        """Конвертує пару в формат для frontend"""
        result = {
            "time": f"{self.begin} - {self.end}",
            "pair_num": self.pair_num
        }
        
        if self.subjects is None:
            result["title"] = "Вільна пара"
            result["type"] = "Вільно"
            result["teacher"] = "-"
            result["room"] = "-"
            result["conference"] = ""
        elif isinstance(self.subjects, Subject):
            result["title"] = self.subjects.subject
            result["type"] = "Лекція"
            result["teacher"] = self.subjects.teacher
            result["room"] = "Не вказано"
            result["conference"] = ""
        elif isinstance(self.subjects, list):
            first_subject = self.subjects[0]
            result["title"] = first_subject.subject + " (1 тиждень)"
            result["type"] = "Лекція"
            result["teacher"] = first_subject.teacher
            result["room"] = "Не вказано"
            result["conference"] = ""
            
            if len(self.subjects) > 1:
                second_subject = self.subjects[1]
                result["title"] += f" / {second_subject.subject} (2 тиждень)"
        
        return result

# Константи
FIRST_GROUP_COLUMN = 4
GROUP_ROW = 2
FIRST_DAY_ROW = 4
ROWS_PER_DAY = 26
DAY_COLUMN = 1
TIME_COLUMN = 2
PAIR_NUMBER_COLUMN = 3

def getRowSize(ws, row, col):
    for mergedRange in ws.merged_cells:
        if mergedRange.min_row <= row <= mergedRange.max_row and mergedRange.min_col <= col <= mergedRange.max_col:
            return mergedRange.max_row - mergedRange.min_row + 1
    return 1

def getCellValue(ws, row, col):
    return ws.cell(row=row, column=col).value

def getFirstDayRow(row):
    return FIRST_DAY_ROW + ((row - FIRST_DAY_ROW) // ROWS_PER_DAY) * ROWS_PER_DAY

def getDay(ws, row):
    firstDayRow = getFirstDayRow(row)
    day = getCellValue(ws, firstDayRow, DAY_COLUMN)
    day_map = {
        'понеділок': 'Понеділок',
        'вівторок': 'Вівторок', 
        'середа': 'Середа',
        'четвер': 'Четвер',
        'п\'ятниця': 'П\'ятниця'
    }
    return day_map.get(day.lower() if day else '', day)

def getPairNumber(ws, row):
    return ws.cell(row=row, column=PAIR_NUMBER_COLUMN).value

def parsePair(ws, row, column):
    day = getDay(ws, row)
    pairNumber = getPairNumber(ws, row)
    
    # Получаем время
    begin = getCellValue(ws, row, TIME_COLUMN)
    if not begin:
        begin = "00:00"
    
    pairColumns = getRowSize(ws, row, PAIR_NUMBER_COLUMN)
    end = getCellValue(ws, row + pairColumns - 1, TIME_COLUMN)
    if not end:
        end = "00:00"
    
    # ОТЛАДКА: Показываем что в ячейке
    subject_value = getCellValue(ws, row, column)
    
    # Пробуем также прочитать напрямую без getCellValue
    direct_value = ws.cell(row=row, column=column).value
    
    # Если это первая пара понедельника - выводим отладку
    if pairNumber == 1 and day == "Понеділок":
        print(f"🔍 DEBUG для строки {row}, столбца {column}:")
        print(f"   getCellValue: {subject_value}")
        print(f"   direct_value: {direct_value}")
        print(f"   cell type: {type(direct_value)}")
        print(f"   is merged: размер ячейки = {getRowSize(ws, row, column)}")
    
    if not subject_value or not str(subject_value).strip():
        # Проверяем может данные в объединенной ячейке
        merged_value = None
        for merged_range in ws.merged_cells:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= column <= merged_range.max_col:
                # Читаем из первой ячейки объединенного диапазона
                merged_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                if pairNumber == 1 and day == "Понеділок":
                    print(f"   merged_value: {merged_value}")
                break
        
        if merged_value and str(merged_value).strip():
            subject_value = merged_value
        else:
            return Pair(day, pairNumber, str(begin), str(end), None)
    
    # Определяем преподавателя
    shift = 2 + pairColumns % 2
    teacher_value = getCellValue(ws, row + shift, column)
    
    # Если не нашли, ищем в следующей строке
    if not teacher_value:
        teacher_value = getCellValue(ws, row + 1, column)
    
    teacher = str(teacher_value).strip() if teacher_value else "Не вказано"
    subject = str(subject_value).strip()
    
    return Pair(day, pairNumber, str(begin), str(end), Subject(subject, teacher))

def parseDay(ws, row, column):
    firstDayRow = getFirstDayRow(row)
    pairs = []
    while row - firstDayRow < ROWS_PER_DAY - 1:
        pair = parsePair(ws, row, column)
        pairColumns = getRowSize(ws, row, PAIR_NUMBER_COLUMN)
        row += pairColumns
        pairs.append(pair)
    return pairs

def parseDays(ws, column):
    row = FIRST_DAY_ROW
    pairs = []
    while getCellValue(ws, row, DAY_COLUMN):
        pairs += parseDay(ws, row, column)
        row += ROWS_PER_DAY
    return pairs

def extractGroups(ws):
    groups = []
    column = FIRST_GROUP_COLUMN
    while True:
        groupName = ws.cell(row=GROUP_ROW, column=column).value
        if not groupName:
            break
        groups.append(groupName)
        column += 2
    return groups

def getGroupColumn(groups: list, groupName: str):
    return FIRST_GROUP_COLUMN + 2 * groups.index(groupName)

def parse_schedule_file(file_path: str, column: int = 57):
    """
    Парсить Excel файл і повертає розклад у форматі для frontend
    
    Args:
        file_path: Шлях до Excel файлу
        column: Номер стовпця для парсингу (за замовчуванням 57 для ОПК-412)
    
    Returns:
        list: Розклад по днях [[Пн], [Вт], [Ср], [Чт], [Пт]]
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    pairs = parseDays(ws, column)
    
    # Групуємо по днях
    schedule_by_day = {
        "Понеділок": [],
        "Вівторок": [],
        "Середа": [],
        "Четвер": [],
        "П'ятниця": []
    }
    
    for pair in pairs:
        if pair.subjects is not None:
            day = pair.day
            if day in schedule_by_day:
                schedule_by_day[day].append(pair.to_dict())
    
    # Конвертуємо в масив
    days_order = ["Понеділок", "Вівторок", "Середа", "Четвер", "П'ятниця"]
    schedule_array = [schedule_by_day[day] for day in days_order]
    
    return schedule_array

# Тест
if __name__ == "__main__":
    print("Тестування парсера...")
    schedule = parse_schedule_file("ОПК-412.xlsx", 57)
    print(json.dumps(schedule, ensure_ascii=False, indent=2))
